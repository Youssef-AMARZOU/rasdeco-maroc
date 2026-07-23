"""
utils_xls.py -- Utilitaires pour parser les fichiers Excel HCP en format cross-table.

Les fichiers HCP sur data.gov.ma ont deux formats :
  - Trimestriel : colonnes = "2006T1", "2006T2"... (i_1.20, i_12.23, etc.)
  - Annuel      : colonnes = "2020", "2021"... (i_1.1, i_1.10, etc.)

Chaque fichier a aussi une feuille "Metadata" avec le nom de l'indicateur.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import polars as pl


def _read_metadata(fpath: Path) -> dict[str, str]:
    """Lit la feuille Metadata d'un fichier HCP et retourne les metadonnees."""
    import openpyxl

    meta = {}
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if "Metadata" not in wb.sheetnames:
            wb.close()
            return meta
        ws = wb["Metadata"]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] and row[1]:
                key = str(row[0]).strip().lower()
                val = str(row[1]).strip()
                if "indicateur" in key:
                    meta["indicateur_label"] = val
                elif "periodicit" in key:
                    meta["periodicite"] = val
                elif "unit" in key:
                    meta["unite"] = val
                elif "source" in key:
                    meta["source"] = val
        wb.close()
    except Exception:
        pass
    return meta


def _detect_indicator_code_from_meta(meta: dict, filename: str) -> str:
    """Detecte le code indicateur a partir des metadonnees HCP et du nom de fichier."""
    label = meta.get("indicateur_label", "").lower()
    periodicite = meta.get("periodicite", "").lower()

    # Mapping motifs -> codes
    if "chomage" in label or "chômege" in label:
        if "trimestr" in periodicite:
            return "CHOMAGE.TRIM"
        return "CHOMAGE.TAUX"
    if "pib" in label or "produit int" in label:
        if "trimestr" in periodicite:
            return "PIB.TRIM.VOL"
        return "PIB.ANNUEL.VOL"
    if "inflation" in label or "ipc" in label or "prix" in label:
        return "IPC.INDICE"
    if "valeur ajout" in label:
        return "VAB.SERVICES"
    if "emploi" in label or "actif" in label:
        return "EMPLOI.VOLUME"
    if "taux de change" in label or "change" in label:
        return "CHANGE.USD"
    if "import" in label:
        return "IMPORTATIONS"
    if "export" in label:
        return "EXPORTATIONS"
    if "dette" in label:
        return "DETTE.PUBLIQUE"
    if "budget" in label or "depense" in label or "recette" in label:
        return "DEFICIT.BUDGET"
    if "tourisme" in label:
        return "TOURISME.ARRIVEES"
    if "agriculture" in label or "agricole" in label or "recolte" in label:
        return "AGRICULTURE.PRODUCTION"

    # Fallback: chercher dans le nom de fichier
    fname = Path(filename).stem.lower()
    if "chomage" in fname:
        return "CHOMAGE.TAUX"
    if "pib" in fname:
        return "PIB.TRIM.VOL"
    if "ipc" in fname or "inflation" in fname:
        return "IPC.INDICE"

    return "INDICATEUR.HCP.GENERIQUE"


def parse_cross_table_xlsx(fpath: Path) -> Optional[pl.DataFrame]:
    """
    Ouvre un fichier XLSX HCP en format cross-table et le convertit en tidy.

    Gere les deux formats :
      - Trimestriel : "2006T1", "2006T2"...
      - Annuel      : "2020", "2021"...

    Retourne un DataFrame avec colonnes : indicateur, date, valeur, code_indicateur, unite.
    """
    import openpyxl

    meta = _read_metadata(fpath)
    indicator_code = _detect_indicator_code_from_meta(meta, str(fpath))

    # Patterns de date : trimestriel ET annuel
    quarterly_pattern = re.compile(r"(20\d{2})\s*[TQ]\s*(\d)")
    annual_pattern = re.compile(r"^(?:\s*)(20\d{2})(?:\s*)$")

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    for ws in wb.worksheets:
        if ws.title.lower() == "metadata":
            continue

        rows = list(ws.iter_rows(values_only=True))

        # Trouver la ligne d'en-tete
        header_row_idx = None
        header_values = None
        is_quarterly = False

        for i, row in enumerate(rows):
            vals = [str(v).strip().replace("\xa0", "") if v is not None else "" for v in row]

            # Compter les motifs trimestriels
            q_count = sum(1 for v in vals if quarterly_pattern.search(v))
            # Compter les motifs annuels (colonnes qui sont juste une annee)
            a_count = sum(
                1 for v in vals
                if annual_pattern.match(v) and not quarterly_pattern.search(v)
            )

            if q_count >= 3:
                header_row_idx = i
                header_values = vals
                is_quarterly = True
                break
            elif a_count >= 3:
                header_row_idx = i
                header_values = vals
                is_quarterly = False
                break

        if header_row_idx is None:
            continue

        # Extraire les dates de l'en-tete
        dates = []
        date_cols = []
        for j, v in enumerate(header_values):
            m = quarterly_pattern.search(v)
            if m:
                year = int(m.group(1))
                quarter = int(m.group(2))
                date_str = f"{year}-{quarter * 3 - 2:02d}-01"
                dates.append(date_str)
                date_cols.append(j)
            elif is_quarterly is False:
                m2 = annual_pattern.match(v)
                if m2:
                    year = int(m2.group(1))
                    dates.append(f"{year}-01-01")
                    date_cols.append(j)

        if not dates:
            continue

        # Lire les lignes de donnees
        unite = meta.get("unite", "%")
        records = []
        for row in rows[header_row_idx + 1:]:
            if not row or all(v is None for v in row):
                continue
            label = str(row[0]).strip() if row[0] is not None else ""
            if not label or len(label) > 80:
                continue
            # Ignorer les lignes de total ou separateurs
            if label.lower() in ("total", "source", "note", "notes", ""):
                continue

            for ci, dj in enumerate(date_cols):
                if dj >= len(row) or row[dj] is None:
                    continue
                raw_val = str(row[dj]).strip().replace("\xa0", "").replace(" ", "")
                if raw_val in ("", "-", "n.d.", "nd", ".."):
                    continue
                try:
                    val = float(raw_val.replace(",", "."))
                    records.append({
                        "indicateur": label,
                        "date": dates[ci],
                        "valeur": val,
                        "code_indicateur": indicator_code,
                        "unite": unite,
                    })
                except (ValueError, TypeError):
                    pass

        if records:
            wb.close()
            return pl.DataFrame(records)

    wb.close()
    return None


def parse_oecd_xls(fpath: Path, skip_rows: int = 3) -> Optional[pl.DataFrame]:
    """
    Parse les fichiers .xls ancien format de l'OCDE / finances.
    """
    try:
        import xlrd
    except ImportError:
        return None

    wb = xlrd.open_workbook(str(fpath))
    sheet = wb.sheet_by_index(0)

    rows = []
    for r in range(sheet.nrows):
        rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])

    # Detection de l'en-tete
    header_idx = None
    for i, row in enumerate(rows):
        date_count = sum(1 for v in row if isinstance(v, str) and re.search(r"20\d{2}", v))
        if date_count >= 3:
            header_idx = i
            break

    if header_idx is None:
        return None

    header = rows[header_idx]
    date_cols = []
    dates = []
    for j, v in enumerate(header):
        v_str = str(v).strip()
        m = re.search(r"(20\d{2})", v_str)
        if m:
            year = int(m.group(1))
            date_str = f"{year}-01-01"
            dates.append(date_str)
            date_cols.append(j)

    records = []
    for row in rows[header_idx + 1:]:
        label = str(row[0]).strip() if row[0] else ""
        if not label or len(label) > 100:
            continue
        for ci, dj in enumerate(date_cols):
            if dj < len(row):
                val = row[dj]
                if isinstance(val, (int, float)):
                    records.append({
                        "indicateur": label,
                        "date": dates[ci] if ci < len(dates) else "2000-01-01",
                        "valeur": float(val),
                    })

    if not records:
        return None
    return pl.DataFrame(records)
