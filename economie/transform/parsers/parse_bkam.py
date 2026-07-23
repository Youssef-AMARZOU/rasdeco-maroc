"""
parsers/parse_bkam.py -- Parser pour les fichiers Bank Al-Maghrib.

Format typique BAM :
- Feuil1: en-tetes multi-lignes, dates (datetime) en rangee 4, indicateurs en rangee 5+
- Plusieurs sections par feuille ("Encours en MDH", "Variation")
- Fichiers XLSX dans des sous-dossiers (ex: 2005/)
- 225+ colonnes de dates (mensuelles, Jan 2005 - Sept 2023)
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl
import polars as pl

from .base import (
    RAW_ROOT,
    ParsedFile,
    SourceParser,
    detect_code_indicateur,
    detect_version_serie,
)

# Mapping des noms d'indicateurs BAM vers codes normalises
_BAM_INDICATOR_MAP = {
    "actif": "BAM.ACTIF",
    "depots": "BAM.DEPOTS",
    "depots aupres des banques": "BAM.DEPOTS",
    "credits": "BAM.CREDITS",
    "titres": "BAM.TITRES",
    "passif": "BAM.PASSIF",
    "capitaux propres": "BAM.CAPITAUX",
    "depots des clients": "BAM.DEPOTS_CLIENTS",
    "emissions": "BAM.EMISSIONS",
    "titres autres qu'actions": "BAM.TITRES_AUTRES",
    "valeur recue en pension": "BAM.PENSION",
    "valeurs recues en pension des autres societes financieres": "BAM.PENSION_SF",
    "creances sur les administrations publiques": "BAM.CREANCES_PAP",
    "obligations du tresor": "BAM.OBLIG_TRESOR",
    "credits aux administrations publiques": "BAM.CREDITS_PAP",
    "autres credits": "BAM.AUTRES_CREDITS",
    "credits hipothecaires": "BAM.CREDITS_HYP",
    "credits a la consommation": "BAM.CREDITS_CONSO",
    "credits aux entreprises": "BAM.CREDITS_ENT",
    "credits a l'economie": "BAM.CREDITS_ECON",
    "creances nettes sur l'exterieur": "BAM.CREANCES_EXT",
    "encaissements": "BAM.ENCAISSEMENTS",
    "actifs externes nets": "BAM.ACTIFS_EXT",
    "monnaie en circulation": "BAM.MONNAIE_CIRC",
    "depots a vue": "BAM.DEPOTS_VUE",
    "depots a terme": "BAM.DEPOTS_TERME",
    "comptes d'epargne": "BAM.COMPTE_EPARGNE",
    "autres depots": "BAM.AUTRES_DEPOTS",
    # OPCVM
    "encours": "BAM.OPCVM.ENCOURS",
    "encours en mdh": "BAM.OPCVM.ENCOURS",
    "actif net": "BAM.OPCVM.ACTIF_NET",
    "nombre de shares": "BAM.OPCVM.SHARES",
    "valeur liquidative": "BAM.OPCVM.VL",
}


def _detect_indicator_code(section: str, label: str) -> str:
    """Detecte le code indicateur a partir de la section et du libelle."""
    combined = f"{section} {label}".lower().strip()
    for kw, code in _BAM_INDICATOR_MAP.items():
        if kw in combined:
            return code
    return detect_code_indicateur(combined) or "BAM.OPCVM.ENCOURS"


def _parse_bam_xlsx(fpath: Path, rel: str, parser: "BKAMParser") -> list[dict]:
    """Parse un fichier BAM XLSX avec format croise (dates en colonnes)."""
    wb = openpyxl.load_workbook(str(fpath), data_only=True, read_only=True)
    rows_out = []

    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 5:
            continue

        # Lire toute la matrice en memoire
        matrix = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            matrix.append(list(row))

        if not matrix:
            continue

        # Trouver la rangee de dates (contient des datetime)
        date_row_idx = None
        for ri in range(min(8, len(matrix))):
            date_count = 0
            for ci in range(1, len(matrix[ri])):
                v = matrix[ri][ci]
                if isinstance(v, datetime):
                    date_count += 1
            if date_count >= 3:
                date_row_idx = ri
                break

        if date_row_idx is None:
            continue

        # Extraire mapping col_index -> date_label
        date_cols = {}
        for ci in range(1, len(matrix[date_row_idx])):
            v = matrix[date_row_idx][ci]
            if isinstance(v, datetime):
                date_cols[ci] = v.strftime("%Y-%m")
            elif v is not None:
                s = str(v).strip()
                if re.match(r"\d{4}", s):
                    date_cols[ci] = s

        if not date_cols:
            continue

        # Extraire indicateurs et valeurs
        current_section = ""
        for ri in range(date_row_idx + 1, len(matrix)):
            label_cell = matrix[ri][0] if len(matrix[ri]) > 0 else None
            if label_cell is None:
                continue
            label = str(label_cell).strip()
            if not label or label.lower() in ("nan", "", "total", "source", "note"):
                continue

            # Verifier si c'est une ligne de donnees (au moins 1 valeur numerique)
            has_numeric = False
            for ci in date_cols:
                if ci < len(matrix[ri]):
                    v = matrix[ri][ci]
                    if v is not None:
                        try:
                            float(str(v).replace(",", ".").replace(" ", ""))
                            has_numeric = True
                            break
                        except (ValueError, TypeError):
                            pass

            if not has_numeric:
                # Titre de section
                current_section = label
                continue

            code = _detect_indicator_code(current_section, label)

            for ci, date_label in date_cols.items():
                if ci >= len(matrix[ri]):
                    continue
                raw_val = matrix[ri][ci]
                if raw_val is None:
                    continue
                try:
                    val = float(str(raw_val).replace(",", ".").replace(" ", ""))
                except (ValueError, TypeError):
                    continue

                rows_out.append(
                    parser._make_row(
                        date_label=date_label,
                        valeur=val,
                        code_indicateur=code,
                        fichier=rel,
                    )
                )

    wb.close()
    return rows_out


class BKAMParser(SourceParser):
    def __init__(self):
        super().__init__("bkam")

    @property
    def source_code(self) -> str:
        return "BAM"

    def parse_file(self, fpath: Path) -> ParsedFile:
        rel = str(fpath.relative_to(RAW_ROOT))
        ext = fpath.suffix.lower()

        if ext == ".pdf":
            return ParsedFile(rel, erreur="PDF ignore (non structure)")
        if ext not in (".xlsx", ".xls", ".csv", ".json"):
            return ParsedFile(rel, erreur=f"Format non supporte : {ext}")

        try:
            if ext in (".xlsx", ".xls"):
                rows = _parse_bam_xlsx(fpath, rel, self)
                if not rows:
                    return ParsedFile(rel, erreur="Aucune serie extraite du fichier Excel")
                return ParsedFile(rel, lignes=pl.DataFrame(rows))
            elif ext == ".csv":
                return self._parse_csv(fpath, rel)
            elif ext == ".json":
                return self._parse_json(fpath, rel)
        except Exception as e:
            return ParsedFile(rel, erreur=str(e))

        return ParsedFile(rel, erreur="Format non gere")

    def _parse_csv(self, fpath: Path, rel: str) -> ParsedFile:
        import pandas as pd

        df = pd.read_csv(fpath, na_values=["", "NA", "..", "n.d."])
        code = detect_code_indicateur(rel) or "TAUX.DIRECTEUR"
        version = detect_version_serie(rel)
        cols = list(df.columns)

        date_col = None
        for c in cols:
            cl = str(c).lower()
            if any(kw in cl for kw in ["date", "annee", "mois", "trimestre", "periode"]):
                date_col = c
                break
        if date_col is None:
            date_col = cols[0]

        value_cols = [c for c in cols if c != date_col and pd.api.types.is_numeric_dtype(df[c])]
        rows = []
        for vcol in value_cols:
            for _, row in df.iterrows():
                raw_val = row[vcol]
                if pd.isna(raw_val):
                    continue
                try:
                    val = float(raw_val)
                except (ValueError, TypeError):
                    continue
                rows.append(
                    self._make_row(
                        date_label=str(row[date_col]),
                        valeur=val,
                        code_indicateur=code,
                        fichier=rel,
                        version_serie=version,
                    )
                )
        if not rows:
            return ParsedFile(rel, erreur="Aucune serie CSV extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    def _parse_json(self, fpath: Path, rel: str) -> ParsedFile:
        import json

        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)

        code = detect_code_indicateur(rel) or "TAUX.DIRECTEUR"
        version = detect_version_serie(rel)
        rows = []

        if isinstance(data, list):
            for item in data:
                if not isinstance(item, dict):
                    continue
                date_val = item.get("date") or item.get("annee") or item.get("periode")
                value = item.get("valeur") or item.get("value") or item.get("montant")
                if date_val and value is not None:
                    try:
                        rows.append(
                            self._make_row(
                                date_label=str(date_val),
                                valeur=float(value),
                                code_indicateur=code,
                                fichier=rel,
                                version_serie=version,
                            )
                        )
                    except (ValueError, TypeError):
                        pass

        if not rows:
            return ParsedFile(rel, erreur="Aucune serie JSON extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))
