"""
parsers/parse_hcp.py -- Parser pour les fichiers HCP (CKAN + BDS).

Analyse les fichiers XLSX telecharges depuis data.gov.ma (org HCP)
ou depuis le BDS. Detecte automatiquement :
  - Series PIB (trimestriel / annuel)
  - IPC / inflation
  - Emploi / chomage
  - Comptes nationaux
"""

from __future__ import annotations

import re
from pathlib import Path

import polars as pl

from .base import (
    INDICATOR_CODES,
    ParsedFile,
    SourceParser,
    _log,
    date_from_label,
    detect_code_indicateur,
    detect_version_serie,
)
from .utils_xls import parse_cross_table_xlsx


class HCParser(SourceParser):
    def __init__(self):
        super().__init__("hcp")

    @property
    def source_code(self) -> str:
        return "HCP"

    def parse_file(self, fpath: Path) -> ParsedFile:
        fpath = fpath.resolve()
        rel = str(fpath.relative_to(Path(__file__).resolve().parents[3] / "data" / "raw" / "economie"))
        ext = fpath.suffix.lower()
        try:
            if ext == ".xlsx":
                return self._parse_xlsx(fpath, rel)
            elif ext == ".csv":
                return self._parse_csv(fpath, rel)
            elif ext == ".json":
                return self._parse_json(fpath, rel)
            else:
                return ParsedFile(rel, erreur=f"Format non supporte : {ext}")
        except Exception as e:
            return ParsedFile(rel, erreur=str(e))

    # ------------------------------------------------------------------
    # XLSX (format principal des donnees HCP)
    # ------------------------------------------------------------------
    def _parse_xlsx(self, fpath: Path, rel: str) -> ParsedFile:
        # Essai 1 : format cross-table HCP (tableau croise avec dates en colonnes)
        try:
            tidy = parse_cross_table_xlsx(fpath)
            if tidy is not None:
                code = detect_code_indicateur(str(fpath)) or "PIB.TRIM.VOL"
                version = detect_version_serie(str(fpath))
                meta = INDICATOR_CODES.get(code, {})
                rows = []
                for row in tidy.iter_rows(named=True):
                    rows.append(self._make_row(
                        date_label=str(row["date"]),
                        valeur=row["valeur"],
                        code_indicateur=code,
                        unite=meta.get("unite", "?"),
                        region_code="MA00",
                        fichier=rel,
                        version_serie=version,
                    ))
                if rows:
                    return ParsedFile(rel, lignes=pl.DataFrame(rows))
        except Exception:
            pass

        # Essai 2 : lecture brute Polars (format standard tidy)
        try:
            sheets = {}
            for sheet_name in self._get_sheet_names(fpath):
                sheets[sheet_name] = pl.read_excel(fpath, sheet_name=sheet_name)
        except Exception:
            sheets = {"Sheet1": pl.read_excel(fpath)}

        code_candidate = detect_code_indicateur(str(fpath))
        version = detect_version_serie(str(fpath))

        rows = []
        for sheet_name, df in sheets.items():
            df = self._normalize_columns(df)
            if df.is_empty():
                continue

            cols = df.columns

            series_type = self._detect_series_type(df, cols, str(fpath), sheet_name)
            indicator_code = series_type.get("code") or code_candidate

            if indicator_code is None:
                rows.extend(self._fallback_extract(df, rel, version))
                continue

            date_col = self._find_date_column(df, cols)
            if date_col is None:
                continue

            value_cols = [c for c in cols if c != date_col and df[c].dtype in (pl.Float64, pl.Int64, pl.Float32)]

            for vcol in value_cols:
                region_code = self._region_from_colname(vcol, str(fpath))
                unite = series_type.get("unite") or INDICATOR_CODES.get(indicator_code, {}).get("unite", "?")

                for row in df.iter_rows(named=True):
                    raw_date = row[date_col]
                    raw_val = row[vcol]
                    if raw_val is None or (isinstance(raw_val, float) and raw_val != raw_val):
                        continue
                    try:
                        val = float(raw_val)
                    except (ValueError, TypeError):
                        continue
                    rows.append(self._make_row(
                        date_label=str(raw_date),
                        valeur=val,
                        code_indicateur=indicator_code,
                        unite=unite,
                        region_code=region_code,
                        fichier=rel,
                        version_serie=version,
                    ))

        if not rows:
            return ParsedFile(rel, erreur="Aucune ligne extraite")

        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------
    def _parse_csv(self, fpath: Path, rel: str) -> ParsedFile:
        try:
            df = pl.read_csv(fpath, infer_schema_length=1000, null_values=["", "NA", "N/A", ".."])
        except Exception:
            df = pl.read_csv(fpath, infer_schema_length=100, null_values=["", "NA"], encoding="utf8-lossy")
        return self._parse_dataframe(df, rel)

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------
    def _parse_json(self, fpath: Path, rel: str) -> ParsedFile:
        try:
            df = pl.read_json(fpath)
        except Exception:
            return ParsedFile(rel, erreur="JSON non tabulaire")
        return self._parse_dataframe(df, rel)

    # ------------------------------------------------------------------
    # DataFrame generique
    # ------------------------------------------------------------------
    def _parse_dataframe(self, df: pl.DataFrame, rel: str) -> ParsedFile:
        if df.is_empty():
            return ParsedFile(rel, erreur="DataFrame vide")
        df = self._normalize_columns(df)
        code = detect_code_indicateur(rel)
        version = detect_version_serie(rel)
        rows = self._fallback_extract(df, rel, version)
        if not rows:
            return ParsedFile(rel, erreur="echec extraction")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    # ------------------------------------------------------------------
    # Heuristiques
    # ------------------------------------------------------------------
    def _get_sheet_names(self, fpath: Path) -> list[str]:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        return wb.sheetnames

    def _detect_series_type(self, df: pl.DataFrame, cols: list[str], fname: str, sheet: str) -> dict:
        """Retourne le code indicateur detecte."""
        text = f"{fname} {sheet} {' '.join(cols)}".lower()
        code = detect_code_indicateur(text)
        meta = INDICATOR_CODES.get(code, {})
        return {"code": code, **meta}

    def _find_date_column(self, df: pl.DataFrame, cols: list[str]) -> str | None:
        """Trouve la colonne qui contient des dates/annees."""
        date_keywords = ["annee", "annee", "annee", "trimestre", "mois", "date", "periode", "periode", "year", "month", "quarter"]
        for col in cols:
            cl = col.lower()
            if any(kw in cl for kw in date_keywords):
                return col

        # Fallback : premiere colonne qui ressemble a des annees
        for col in cols:
            sample = df[col].head(5).to_list()
            if all(isinstance(v, (int, float)) and 1950 <= v <= 2030 for v in sample if v is not None):
                return col
        return None

    def _region_from_colname(self, col: str, fname: str) -> str:
        """Extrait un code region depuis le nom de colonne ou fichier."""
        cl = col.lower()
        text = f"{cl} {fname.lower()}"
        region_map = {
            "national": "MA00",
            "maroc": "MA00",
            "tanger": "MA01",
            "tetouan": "MA01",
            "oriental": "MA02",
            "fes": "MA03",
            "fes": "MA03",
            "meknes": "MA03",
            "meknes": "MA03",
            "rabat": "MA04",
            "kenitra": "MA04",
            "beni mellal": "MA05",
            "beni mellal": "MA05",
            "khenifra": "MA05",
            "khenifra": "MA05",
            "casablanca": "MA06",
            "settat": "MA06",
            "marrakech": "MA07",
            "safi": "MA07",
            "drâa": "MA08",
            "dara": "MA08",
            "souss": "MA09",
            "agadir": "MA09",
            "guelmim": "MA10",
            "laâyoune": "MA11",
            "laayoune": "MA11",
            "dakhla": "MA12",
        }
        for kw, code in region_map.items():
            if kw in text:
                return code
        return "MA00"

    def _fallback_extract(self, df: pl.DataFrame, rel: str, version: str) -> list[dict]:
        """
        Extraction generique : colonne de gauche = dates,
        colonnes numeriques = series.
        """
        df = self._normalize_columns(df)
        if df.is_empty():
            return []

        cols = df.columns
        date_col = self._find_date_column(df, cols)
        if date_col is None:
            return []

        value_cols = [c for c in cols if c != date_col and df[c].dtype in (pl.Float64, pl.Int64, pl.Float32, pl.Int32)]
        if not value_cols:
            return []

        code = detect_code_indicateur(rel) or "PIB.TRIM.VOL"
        rows = []
        for vcol in value_cols:
            region = self._region_from_colname(vcol, rel)
            for row in df.iter_rows(named=True):
                raw_date = row[date_col]
                raw_val = row[vcol]
                if raw_val is None or (isinstance(raw_val, float) and raw_val != raw_val):
                    continue
                try:
                    val = float(raw_val)
                except (ValueError, TypeError):
                    continue
                rows.append(self._make_row(
                    date_label=str(raw_date),
                    valeur=val,
                    code_indicateur=code,
                    region_code=region,
                    fichier=rel,
                    version_serie=version,
                ))
        return rows
