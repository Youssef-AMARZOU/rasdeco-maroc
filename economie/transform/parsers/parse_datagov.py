"""
parsers/parse_datagov.py -- Parser pour les fichiers data.gov.ma (tous organismes).

Ce parser est le plus generique : il traite les fichiers issus du
groupe « economie et Finance » via le CKAN de data.gov.ma,
quelle que soit l'organisation source.
"""

from __future__ import annotations

import re
from pathlib import Path

import polars as pl

from .base import (
    INDICATOR_CODES,
    ParsedFile,
    SourceParser,
    _log,
    detect_code_indicateur,
    detect_version_serie,
)


class DatagovParser(SourceParser):
    def __init__(self):
        super().__init__("datagov")

    @property
    def source_code(self) -> str:
        return "DATAGOV"

    def parse_file(self, fpath: Path) -> ParsedFile:
        rel = str(fpath.relative_to(
            Path(__file__).resolve().parents[3] / "data" / "raw" / "economie"
        ))
        ext = fpath.suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                return self._parse_xlsx(fpath, rel)
            elif ext == ".csv":
                return self._parse_csv(fpath, rel)
            elif ext == ".json":
                return self._parse_json(fpath, rel)
            elif ext == ".pdf":
                return ParsedFile(rel, erreur="PDF ignore (non structure)")
            elif ext == ".xml":
                return self._parse_xml(fpath, rel)
            else:
                return ParsedFile(rel, erreur=f"Format non supporte : {ext}")
        except Exception as e:
            return ParsedFile(rel, erreur=str(e))

    # ------------------------------------------------------------------
    # XLSX generique data.gov.ma
    # ------------------------------------------------------------------
    def _parse_xlsx(self, fpath: Path, rel: str) -> ParsedFile:
        rows = []
        code = detect_code_indicateur(rel)
        version = detect_version_serie(rel)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
        except Exception:
            sheet_names = ["Sheet1"]

        for sheet in sheet_names:
            try:
                df = pl.read_excel(fpath, sheet_name=sheet)
            except Exception:
                continue
            if df.is_empty():
                continue
            rows.extend(self._extract_generic(df, rel, code, version))
            if rows:
                break

        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    # ------------------------------------------------------------------
    # CSV generique
    # ------------------------------------------------------------------
    def _parse_csv(self, fpath: Path, rel: str) -> ParsedFile:
        encodings = ["utf-8", "latin-1", "iso-8859-1"]
        df = None
        for enc in encodings:
            try:
                df = pl.read_csv(fpath, infer_schema_length=500, null_values=["", "NA", ".."], encoding=enc)
                break
            except Exception:
                continue
        if df is None or df.is_empty():
            return ParsedFile(rel, erreur="echec lecture CSV")
        code = detect_code_indicateur(rel)
        version = detect_version_serie(rel)
        rows = self._extract_generic(df, rel, code, version)
        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------
    def _parse_json(self, fpath: Path, rel: str) -> ParsedFile:
        try:
            df = pl.read_json(fpath)
        except Exception:
            return ParsedFile(rel, erreur="JSON non tabulaire")
        code = detect_code_indicateur(rel)
        version = detect_version_serie(rel)
        rows = self._extract_generic(df, rel, code, version)
        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    # ------------------------------------------------------------------
    # XML (via ElementTree)
    # ------------------------------------------------------------------
    def _parse_xml(self, fpath: Path, rel: str) -> ParsedFile:
        import xml.etree.ElementTree as ET
        tree = ET.parse(fpath)
        root = tree.getroot()
        data_records = []
        for child in root.iter():
            if child.tag == "record" or child.tag == "row":
                rec = {sub.tag: sub.text for sub in child}
                data_records.append(rec)
        if not data_records:
            return ParsedFile(rel, erreur="XML sans records")
        df = pl.DataFrame(data_records)
        code = detect_code_indicateur(rel)
        version = detect_version_serie(rel)
        rows = self._extract_generic(df, rel, code, version)
        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite du XML")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    # ------------------------------------------------------------------
    # Extraction generique (fallback universel)
    # ------------------------------------------------------------------
    def _extract_generic(self, df: pl.DataFrame, rel: str, code_hint: str | None, version: str) -> list[dict]:
        df = self._normalize_columns(df)
        if df.is_empty():
            return []

        # Detection multi-indicateurs : si le DataFrame contient
        # une colonne "indicateur", "variable", etc. on pivot.
        indicator_col = None
        for c in df.columns:
            cl = c.lower()
            if any(kw in cl for kw in ["indicateur", "variable", "serie", "serie", "indicator"]):
                indicator_col = c
                break

        date_col = None
        for c in df.columns:
            cl = c.lower()
            if any(kw in cl for kw in ["annee", "annee", "mois", "trimestre", "date", "periode", "periode", "year", "month"]):
                date_col = c
                break

        if date_col is None:
            # Fallback : premiere colonne
            date_col = df.columns[0]

        rows = []

        if indicator_col and date_col:
            # Format long : plusieurs indicateurs dans la meme colonne
            value_cols = [c for c in df.columns if c not in (date_col, indicator_col) and df[c].dtype in (pl.Float64, pl.Int64, pl.Float32)]
            for vcol in value_cols:
                for row in df.iter_rows(named=True):
                    ind_code_raw = str(row[indicator_col]) if row[indicator_col] else ""
                    ind_code = detect_code_indicateur(ind_code_raw) or code_hint or "PIB.TRIM.VOL"
                    raw_date = row[date_col]
                    raw_val = row[vcol]
                    if raw_val is None or (isinstance(raw_val, float) and raw_val != raw_val):
                        continue
                    try:
                        val = float(raw_val)
                    except (ValueError, TypeError):
                        continue
                    rows.append(self._make_row(
                        date_label=str(raw_date),
                        valeur=val,
                        code_indicateur=ind_code,
                        fichier=rel,
                        version_serie=version,
                    ))
        else:
            # Format large : cols = dates, lignes = indicateurs, ou l'inverse
            code = code_hint or detect_code_indicateur(rel) or "PIB.TRIM.VOL"
            value_cols = [c for c in df.columns if c != date_col and df[c].dtype in (pl.Float64, pl.Int64, pl.Float32)]
            for vcol in value_cols:
                for row in df.iter_rows(named=True):
                    raw_date = row[date_col]
                    raw_val = row[vcol]
                    if raw_val is None or (isinstance(raw_val, float) and raw_val != raw_val):
                        continue
                    try:
                        val = float(raw_val)
                    except (ValueError, TypeError):
                        continue
                    rows.append(self._make_row(
                        date_label=str(raw_date),
                        valeur=val,
                        code_indicateur=vcol if code_hint is None else code,
                        fichier=rel,
                        version_serie=version,
                    ))

        return rows
