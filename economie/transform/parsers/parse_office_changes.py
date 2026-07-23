"""
parsers/parse_office_changes.py -- Parser pour l'Office des Changes.

Sources :
  - CKAN data.gov.ma
  - Publications officielles oc.gov.ma (PDF/XLSX)
  - Scraping

Les donnees portent sur : reserves de change, IDE, balance commerciale.
"""

from __future__ import annotations

from pathlib import Path

import polars as pl

from .base import (
    INDICATOR_CODES,
    ParsedFile,
    SourceParser,
    detect_code_indicateur,
    detect_version_serie,
)


class OfficeChangesParser(SourceParser):
    def __init__(self):
        super().__init__("office_des_changes")

    @property
    def source_code(self) -> str:
        return "OC"

    def parse_file(self, fpath: Path) -> ParsedFile:
        rel = str(fpath.relative_to(
            Path(__file__).resolve().parents[3] / "data" / "raw" / "economie"
        ))
        ext = fpath.suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                return self._parse_xlsx(fpath, rel)
            elif ext == ".csv":
                return self._parse_csv(fpath, rel)
            elif ext == ".json":
                return self._parse_json(fpath, rel)
            elif ext == ".pdf":
                return ParsedFile(rel, erreur="PDF ignore (non structure)")
            else:
                return ParsedFile(rel, erreur=f"Format non supporte : {ext}")
        except Exception as e:
            return ParsedFile(rel, erreur=str(e))

    def _parse_xlsx(self, fpath: Path, rel: str) -> ParsedFile:
        code = detect_code_indicateur(rel) or "BALANCE.COMMERCIALE"
        version = detect_version_serie(rel)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            sheets = wb.sheetnames
        except Exception:
            sheets = ["Sheet1"]

        rows = []
        for sheet in sheets:
            try:
                df = pl.read_excel(fpath, sheet_name=sheet)
            except Exception:
                continue
            extracted = self._extract(df, rel, code, version)
            rows.extend(extracted)
            if rows:
                break

        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    def _parse_csv(self, fpath: Path, rel: str) -> ParsedFile:
        df = pl.read_csv(fpath, infer_schema_length=1000, null_values=["", "NA", ".."])
        code = detect_code_indicateur(rel) or "BALANCE.COMMERCIALE"
        version = detect_version_serie(rel)
        rows = self._extract(df, rel, code, version)
        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    def _parse_json(self, fpath: Path, rel: str) -> ParsedFile:
        df = pl.read_json(fpath)
        code = detect_code_indicateur(rel) or "BALANCE.COMMERCIALE"
        version = detect_version_serie(rel)
        rows = self._extract(df, rel, code, version)
        if not rows:
            return ParsedFile(rel, erreur="Aucune serie extraite")
        return ParsedFile(rel, lignes=pl.DataFrame(rows))

    def _extract(self, df: pl.DataFrame, rel: str, code: str, version: str) -> list[dict]:
        if df.is_empty():
            return []
        df = self._normalize_columns(df)
        cols = df.columns

        # Trois colonnes-tendances : date, libelle (optionnel), valeurs
        date_col = None
        for c in cols:
            cl = c.lower()
            if any(kw in cl for kw in ["annee", "annee", "mois", "date", "trimestre"]):
                date_col = c
                break
        if date_col is None:
            date_col = cols[0]

        value_cols = [c for c in cols if c != date_col and df[c].dtype in (pl.Float64, pl.Int64, pl.Float32)]
        if not value_cols:
            return []

        rows = []
        for vcol in value_cols:
            for row in df.iter_rows(named=True):
                raw_date = row[date_col]
                raw_val = row[vcol]
                if raw_val is None or (isinstance(raw_val, float) and raw_val != raw_val):
                    continue
                try:
                    val = float(raw_val)
                except (ValueError, TypeError):
                    continue
                rows.append(self._make_row(
                    date_label=str(raw_date),
                    valeur=val,
                    code_indicateur=code,
                    fichier=rel,
                    version_serie=version,
                ))
        return rows
